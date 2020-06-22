import openpyxl
wb = openpyxl.load_workbook("sample.xlsx")
wb.sheetnames
wb.active
sheet = wb['Sheet1']
sheet.title
sheet.cell(row=1, column=2)